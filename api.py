"""
E-Trade Foreign Share Working — FastAPI Backend (Memory Optimized)
Run: uvicorn api:app --reload
"""

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pandas as pd
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from copy import copy
import pdfplumber
import re
import io
import os
import gc
import tempfile
from datetime import datetime
from typing import List, Optional

app = FastAPI(title="E-Trade Foreign Share Working API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FY_MAP = {
    "FY 2023-24": ("2023-04-01", "2024-03-31"),
    "FY 2024-25": ("2024-04-01", "2025-03-31"),
    "FY 2025-26": ("2025-04-01", "2026-03-31"),
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_col(headers, keywords):
    for h_text, col_idx in headers.items():
        if any(k.lower() in str(h_text).lower() for k in keywords):
            return col_idx
    return None

def copy_row_style(ws, src, dst):
    for c in range(1, ws.max_column + 1):
        s_cell = ws.cell(src, c)
        d_cell = ws.cell(dst, c)
        if isinstance(s_cell.value, str) and s_cell.value.startswith("="):
            d_cell.value = Translator(s_cell.value, origin=s_cell.coordinate).translate_formula(d_cell.coordinate)
        else:
            d_cell.value = s_cell.value
        if s_cell.has_style:
            d_cell.font        = copy(s_cell.font)
            d_cell.border      = copy(s_cell.border)
            d_cell.fill        = copy(s_cell.fill)
            d_cell.number_format = s_cell.number_format
            d_cell.alignment   = copy(s_cell.alignment)

def extract_dividends_memory_safe(pdf_bytes_list, start_dt):
    """Process PDFs one page at a time to minimise memory usage."""
    recs = []
    for pdf_bytes in pdf_bytes_list:
        base_yr = start_dt.year
        temp_div = None
        # Write to temp file so pdfplumber doesn't hold full bytes in RAM
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(pdf_bytes)
            tmp_path = tmp.name
        try:
            with pdfplumber.open(tmp_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    for line in text.split("\n"):
                        if "For the Period" in line:
                            m = re.search(r"\b(20\d{2})\b", line)
                            if m:
                                base_yr = int(m.group(1))
                        if "Qualified Dividend" in line:
                            try:
                                d_m = re.search(r"(\d{1,2}/\d{1,2})", line)
                                a_m = re.findall(r"[\d,]+\.\d+", line)
                                if d_m:
                                    mo, day = d_m.group(1).split("/")
                                    script = re.sub(r"[\d,]+\.\d+", "", line.split("Qualified Dividend")[-1]).strip()
                                    temp_div = {
                                        "Date":     datetime(base_yr, int(mo), int(day)),
                                        "Script":   script,
                                        "Dividend": float(a_m[-1].replace(",", "")),
                                        "Tax":      0,
                                    }
                            except Exception:
                                continue
                        elif "Tax Withholding" in line and temp_div:
                            try:
                                temp_div["Tax"] = float(re.findall(r"[\d,]+\.\d+", line)[-1].replace(",", ""))
                                recs.append(temp_div)
                                temp_div = None
                            except Exception:
                                continue
                        elif "Interest Income" in line:
                            try:
                                d_m = re.search(r"(\d{1,2})/(\d{1,2})", line)
                                a_m = re.findall(r"[\d,]+\.\d+", line)
                                if d_m:
                                    recs.append({
                                        "Date":     datetime(base_yr, int(d_m.group(1)), int(d_m.group(2))),
                                        "Script":   "Interest Income",
                                        "Dividend": float(a_m[-1].replace(",", "")),
                                        "Tax":      0,
                                    })
                            except Exception:
                                continue
                    # Free page memory immediately
                    del text
                # Free PDF pages
                del pdf
        finally:
            os.unlink(tmp_path)
            gc.collect()  # Force garbage collection after each PDF

    return pd.DataFrame(recs) if recs else pd.DataFrame(columns=["Date","Script","Dividend","Tax"])


def build_report(template_bytes, gl_bytes_list, pdf_bytes_list, client_name, pan_number, fy):
    if fy not in FY_MAP:
        raise ValueError(f"Unknown FY: {fy}")

    start_dt, end_dt = map(pd.to_datetime, FY_MAP[fy])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(template_bytes)
        tmp_path = tmp.name

    # Free template bytes from memory
    del template_bytes
    gc.collect()

    try:
        wb = openpyxl.load_workbook(tmp_path)

        # ---- CAPITAL GAINS ----
        if gl_bytes_list and "Capital Gains" in wb.sheetnames:
            ws = wb["Capital Gains"]

            frames = []
            for gl_bytes in gl_bytes_list:
                frames.append(pd.read_excel(io.BytesIO(gl_bytes)))
                del gl_bytes
            gc.collect()

            df_cg = pd.concat(frames, ignore_index=True)
            del frames
            gc.collect()

            df_cg.columns = df_cg.columns.str.strip()
            qty_col = next((c for c in df_cg.columns if c in ["Quantity", "Qty", "Qty."]), "Quantity")
            df_cg["Date Sold"]     = pd.to_datetime(df_cg["Date Sold"],     errors="coerce")
            df_cg["Date Acquired"] = pd.to_datetime(df_cg["Date Acquired"], errors="coerce")
            df_f = df_cg[(df_cg["Date Sold"] >= start_dt) & (df_cg["Date Sold"] <= end_dt)].copy()
            del df_cg
            gc.collect()

            h_row = 1
            for r in range(1, 15):
                if any("Script" in str(ws.cell(r, c).value) for c in range(1, 10)):
                    h_row = r; break

            headers = {str(ws.cell(h_row, c).value).strip(): c for c in range(1, ws.max_column+1) if ws.cell(h_row, c).value}
            c_scr = find_col(headers, ["Script"])
            c_qty = find_col(headers, ["Shares", "Qty"])
            c_sd  = find_col(headers, ["Sale Date"])
            c_pd  = find_col(headers, ["Purchase Date"])
            c_sv  = find_col(headers, ["Sale Value"])
            c_pv  = find_col(headers, ["Purchase Value", "Cost"])

            data_start = h_row + 2
            ws.delete_rows(data_start + 1, 1000)

            rows = df_f.to_dict("records")
            del df_f
            gc.collect()

            for i, row in enumerate(rows):
                curr = data_start + i
                copy_row_style(ws, data_start, curr)
                if c_scr: ws.cell(curr, c_scr).value = row.get("Plan Type", "Shares")
                if c_qty: ws.cell(curr, c_qty).value = row.get(qty_col)
                if c_sd:
                    ws.cell(curr, c_sd).value = row.get("Date Sold")
                    ws.cell(curr, c_sd).number_format = "DD-MM-YYYY"
                if c_pd:
                    ws.cell(curr, c_pd).value = row.get("Date Acquired")
                    ws.cell(curr, c_pd).number_format = "DD-MM-YYYY"
                if c_sv: ws.cell(curr, c_sv).value = row.get("Total Proceeds")
                if c_pv: ws.cell(curr, c_pv).value = row.get("Adjusted Cost Basis")

            b_start = data_start + len(rows)
            for b in range(2):
                copy_row_style(ws, data_start, b_start + b)
                for idx in [c_scr, c_qty, c_sd, c_pd, c_sv, c_pv]:
                    if idx: ws.cell(b_start + b, idx).value = None

            total_r = b_start + 2
            if c_scr: ws.cell(total_r, c_scr).value = "Total"
            for h_n, h_idx in headers.items():
                if any(x in h_n.lower() for x in ["shares", "value", "gain", "inr"]):
                    col_l = get_column_letter(h_idx)
                    ws.cell(total_r, h_idx).value = f"=SUM({col_l}{data_start}:{col_l}{total_r-1})"

            del rows
            gc.collect()

        # ---- DIVIDENDS ----
        if pdf_bytes_list and "Dividends" in wb.sheetnames:
            ws = wb["Dividends"]

            df_div = extract_dividends_memory_safe(pdf_bytes_list, start_dt)
            del pdf_bytes_list
            gc.collect()

            df_div = df_div[(df_div["Date"] >= start_dt) & (df_div["Date"] <= end_dt)]

            h_row = 1
            for r in range(1, 15):
                if any("Dividend" in str(ws.cell(r, c).value) for c in range(1, 10)):
                    h_row = r; break

            headers = {str(ws.cell(h_row, c).value).strip(): c for c in range(1, ws.max_column+1) if ws.cell(h_row, c).value}
            d_dt  = find_col(headers, ["Date"])
            d_scr = find_col(headers, ["Script"])
            d_dv  = find_col(headers, ["Dividend"])
            d_tx  = find_col(headers, ["Tax"])

            data_start = h_row + 2
            ws.delete_rows(data_start + 1, 1000)

            rows = df_div.to_dict("records")
            del df_div
            gc.collect()

            for i, row in enumerate(rows):
                curr = data_start + i
                copy_row_style(ws, data_start, curr)
                if d_dt:
                    ws.cell(curr, d_dt).value = row.get("Date")
                    ws.cell(curr, d_dt).number_format = "DD-MM-YYYY"
                if d_scr: ws.cell(curr, d_scr).value = row.get("Script")
                if d_dv:  ws.cell(curr, d_dv).value  = row.get("Dividend")
                if d_tx:  ws.cell(curr, d_tx).value  = row.get("Tax")

            b_start = data_start + len(rows)
            for b in range(2):
                copy_row_style(ws, data_start, b_start + b)
                for idx in [d_dt, d_scr, d_dv, d_tx]:
                    if idx: ws.cell(b_start + b, idx).value = None

            total_r = b_start + 2
            if d_scr: ws.cell(total_r, d_scr).value = "Total"
            for h_n, h_idx in headers.items():
                if any(x in h_n.lower() for x in ["dividend", "tax", "inr"]):
                    col_l = get_column_letter(h_idx)
                    ws.cell(total_r, h_idx).value = f"=SUM({col_l}{data_start}:{col_l}{total_r-1})"

            del rows
            gc.collect()

        # ---- TOP INFO ----
        for sn in ["Capital Gains", "Dividends"]:
            if sn in wb.sheetnames:
                ws = wb[sn]
                for r in range(1, 12):
                    for c in range(1, 5):
                        v = str(ws.cell(r, c).value).lower() if ws.cell(r, c).value else ""
                        if "name"   in v: ws.cell(r, c+1).value = client_name
                        if "pan"    in v: ws.cell(r, c+1).value = pan_number.upper()
                        if "period" in v: ws.cell(r, c+1).value = fy

        out = io.BytesIO()
        wb.save(out)
        del wb
        gc.collect()
        return out.getvalue()

    finally:
        os.unlink(tmp_path)
        gc.collect()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/health")
def health():
    return {"status": "ok", "time": datetime.now().isoformat()}


@app.post("/generate")
async def generate(
    template:    UploadFile       = File(...),
    gl_files:    List[UploadFile] = File([]),
    pdf_files:   List[UploadFile] = File([]),
    client_name: str              = Form(...),
    pan_number:  str              = Form(...),
    fy:          str              = Form(...),
):
    template_bytes  = await template.read()
    gl_bytes_list   = [await f.read() for f in gl_files]
    pdf_bytes_list  = [await f.read() for f in pdf_files]

    try:
        xlsx_bytes = build_report(
            template_bytes  = template_bytes,
            gl_bytes_list   = gl_bytes_list,
            pdf_bytes_list  = pdf_bytes_list,
            client_name     = client_name,
            pan_number      = pan_number,
            fy              = fy,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        gc.collect()

    safe_name = client_name.replace(" ", "_") if client_name else "Client"
    filename  = f"Foreign_Shares_{safe_name}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/preview-gl")
async def preview_gl(
    gl_files: List[UploadFile] = File(...),
    fy:       str              = Form(...),
):
    if fy not in FY_MAP:
        raise HTTPException(status_code=400, detail="Unknown FY")

    start_dt, end_dt = map(pd.to_datetime, FY_MAP[fy])
    frames = []
    for f in gl_files:
        frames.append(pd.read_excel(io.BytesIO(await f.read())))
    df = pd.concat(frames, ignore_index=True)
    del frames
    gc.collect()

    df.columns = df.columns.str.strip()
    df["Date Sold"]     = pd.to_datetime(df["Date Sold"],     errors="coerce")
    df["Date Acquired"] = pd.to_datetime(df["Date Acquired"], errors="coerce")
    df = df[(df["Date Sold"] >= start_dt) & (df["Date Sold"] <= end_dt)]

    for col in ["Date Sold", "Date Acquired"]:
        if col in df.columns:
            df[col] = df[col].dt.strftime("%d-%m-%Y")

    result = {"rows": df.fillna("").to_dict("records"), "count": len(df)}
    del df
    gc.collect()
    return result


@app.post("/preview-dividends")
async def preview_dividends(
    pdf_files: List[UploadFile] = File(...),
    fy:        str              = Form(...),
):
    if fy not in FY_MAP:
        raise HTTPException(status_code=400, detail="Unknown FY")

    start_dt, end_dt = map(pd.to_datetime, FY_MAP[fy])
    pdf_bytes_list = [await f.read() for f in pdf_files]

    df = extract_dividends_memory_safe(pdf_bytes_list, start_dt)
    del pdf_bytes_list
    gc.collect()

    df = df[(df["Date"] >= start_dt) & (df["Date"] <= end_dt)]
    df["Date"] = df["Date"].dt.strftime("%d-%m-%Y")

    result = {"rows": df.fillna("").to_dict("records"), "count": len(df)}
    del df
    gc.collect()
    return result
